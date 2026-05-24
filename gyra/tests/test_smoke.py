"""Smoke tests — confirm the app boots and core routes respond.

Run with either:
    python3 -m unittest tests.test_smoke      # stdlib, no install
    pytest tests/test_smoke.py                # if pytest is installed

Anything beyond a 5xx here means a wiring/import problem at app startup.
"""
from __future__ import annotations

import unittest

from tests.conftest import make_admin_client


class SmokeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        import app as appmod
        cls.app = appmod.app
        cls.client = make_admin_client(cls.app)

    def test_login_page_renders(self):
        # /login is unauthenticated, so this proves the app booted.
        anon = self.app.test_client()
        r = anon.get("/login")
        self.assertEqual(r.status_code, 200)

    def test_dashboard_authed(self):
        r = self.client.get("/")
        # Either the dashboard renders (200) or it redirects to a default
        # project board (302). Both indicate a healthy boot.
        self.assertIn(r.status_code, (200, 302))

    def test_xlsx_template_export(self):
        # Regression guard for the bulk-add Excel export sheet ordering.
        from openpyxl import load_workbook
        import io
        r = self.client.get("/project/3/bulk-add/xlsx-template")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.data))
        self.assertEqual(
            wb.sheetnames,
            ["Stories", "All Epics", "All Initiatives", "Help", "Lists"],
        )


if __name__ == "__main__":
    unittest.main()
