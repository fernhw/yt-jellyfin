"""routes/sprint.py — Bulk story add, end-sprint, archive, and unarchive."""
import csv
import io
import time

from flask import (abort, flash, jsonify, redirect,
                   render_template, request, session, url_for)
from markupsafe import Markup, escape

from auth import enforce_csrf, login_required, super_user_required
from db import (get_all_active_users, get_current_sprint, get_db,
                get_epics, get_initiatives, get_project, get_story_types, user_in_project)
from routes.helpers import (ACTOR_OPTIONS, CONNECTOR_OPTIONS, VERB_OPTIONS,
                            bold_verb_in_title, build_story_title,
                            validate_story_parts)
from db import get_story_previews

# Allowed story-point values (modified Fibonacci). 0 is allowed and means
# "no estimate yet"; anything else outside this set is rejected by the
# bulk-add UI and the CSV importer.
FIB_POINTS = {0, 1, 2, 3, 5, 8, 13, 21}


def _parse_points(raw: str):
    """Return (points:int, error:str|None). Empty → (0, None)."""
    s = (raw or "").strip()
    if not s:
        return 0, None
    try:
        v = int(s)
    except ValueError:
        return 0, "Story points must be a number ({} not valid).".format(s)
    if v not in FIB_POINTS:
        return 0, "Story points must be Fibonacci: {} (got {}).".format(
            ", ".join(str(n) for n in sorted(FIB_POINTS)), v)
    return v, None


def _check_project_access(project_id: int):
    """Return project or abort; also 403 if user not in project."""
    project = get_project(project_id)
    if not project:
        abort(404)
    if session.get("role") != "admin" and not user_in_project(session["user_id"], project_id):
        abort(403)
    return project


def register(app) -> None:

    # ── Bulk story add ────────────────────────────────────────────────────────

    @app.route("/project/<int:project_id>/bulk-add", methods=["GET", "POST"])
    @super_user_required
    def bulk_add(project_id):
        project = _check_project_access(project_id)

        if request.method == "POST":
            enforce_csrf()
            actors       = request.form.getlist("story_actor[]")
            verbs        = request.form.getlist("story_verb[]")
            zs           = request.form.getlist("story_z[]")
            xs           = request.form.getlist("story_x[]")
            fors         = request.form.getlist("story_for[]")
            ys           = request.form.getlist("story_y[]")
            points_list  = request.form.getlist("points[]")
            priorities   = request.form.getlist("priority[]")
            descs        = request.form.getlist("description[]")
            assignees    = request.form.getlist("assignee[]")
            story_types_list = request.form.getlist("story_type[]")
            epic_ids     = request.form.getlist("epic_id[]")

            n = max(len(actors), len(verbs), len(zs), len(xs), len(fors), len(ys))

            def at(lst, i):
                return lst[i].strip() if i < len(lst) else ""

            conn = get_db()
            first_status = conn.execute(
                "SELECT id FROM statuses WHERE project_id=? ORDER BY order_index LIMIT 1",
                (project_id,),
            ).fetchone()
            status_id = first_status["id"] if first_status else None
            now       = int(time.time())
            created   = 0
            skipped   = 0
            row_errors = []   # list[(row_num, [messages])]
            row_warnings = [] # list[(row_num, [messages])]

            for i in range(n):
                actor    = at(actors, i)
                verb     = at(verbs, i)
                z        = at(zs, i)
                x        = at(xs, i)
                for_conn = at(fors, i)
                y        = at(ys, i)

                # Skip totally-empty rows silently (no Action Word, no What,
                # no Outcome — clearly a placeholder row the user never filled in).
                if not (z or x or y):
                    continue

                errors, warnings = validate_story_parts(
                    actor, verb, z, x, for_conn, y,
                    points=(_parse_points(at(points_list, i))[0]
                            if at(points_list, i) else None),
                    description=at(descs, i),
                )
                row_num = i + 1
                if errors:
                    row_errors.append((row_num, errors))
                    skipped += 1
                    continue
                if warnings:
                    row_warnings.append((row_num, warnings))

                title = build_story_title(actor, verb, z, x, for_conn, y)

                pts_raw = at(points_list, i)
                pts, pts_err = _parse_points(pts_raw)
                if pts_err:
                    row_errors.append((row_num, [pts_err]))
                    skipped += 1
                    continue
                prio    = at(priorities, i) or None
                desc    = at(descs, i)
                a_raw   = at(assignees, i)
                a_id    = int(a_raw) if a_raw.isdigit() else None
                st_raw  = at(story_types_list, i)
                st_id   = int(st_raw) if st_raw.isdigit() else None
                ep_raw  = at(epic_ids, i)
                ep_id   = int(ep_raw) if ep_raw.isdigit() else None

                max_idx = conn.execute(
                    "SELECT COALESCE(MAX(order_index), 0) + 1 FROM stories WHERE project_id=?",
                    (project_id,),
                ).fetchone()[0]
                cur = conn.execute(
                    """INSERT INTO stories
                       (project_id, title, description, story_points, priority,
                        story_type, epic_id,
                        status_id, created_at, created_by, order_index, is_archived,
                        story_actor, story_verb, story_z, story_x, story_for, story_y)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,0,?,?,?,?,?,?)""",
                    (project_id, title, desc, pts, prio,
                     st_id, ep_id,
                     status_id, now, session["user_id"], max_idx,
                     actor, verb, z, x, for_conn, y),
                )
                if a_id:
                    conn.execute(
                        "INSERT OR IGNORE INTO story_users (story_id, user_id, role) VALUES (?,?,'assignee')",
                        (cur.lastrowid, a_id),
                    )
                created += 1

            conn.commit()
            conn.close()

            # Build flash messages — keep them readable per-row.
            if row_errors:
                lines = ["<strong>{} row(s) rejected.</strong>".format(skipped)]
                for row_num, errs in row_errors:
                    lines.append("<br><strong>Row {}:</strong>".format(row_num))
                    for e in errs:
                        lines.append("<br>&nbsp;&nbsp;• " + str(escape(e)))
                flash(Markup("".join(lines)), "error")

            if row_warnings:
                wlines = ["<strong>Warnings:</strong>"]
                for row_num, warns in row_warnings:
                    wlines.append("<br><strong>Row {}:</strong>".format(row_num))
                    for w in warns:
                        wlines.append("<br>&nbsp;&nbsp;• " + str(escape(w)))
                flash(Markup("".join(wlines)), "warning")

            if created:
                flash(f"{created} {'story' if created == 1 else 'stories'} created.", "success")
            elif not row_errors:
                flash("Nothing to create — all rows were empty.", "warning")

            # If everything succeeded, send the user to the backlog. Otherwise
            # stay on the bulk-add page so they can fix the rejected rows.
            if created and not row_errors:
                return redirect(url_for("backlog", project_id=project_id))
            return redirect(url_for("bulk_add", project_id=project_id))

        users       = get_all_active_users()
        story_types = get_story_types(project_id)
        epics       = get_epics(project_id)
        prefill_rows = session.pop("bulk_prefill", None)
        return render_template("bulk_add.html", project=project, users=users,
                               story_types=story_types, epics=epics,
                               prefill_rows=prefill_rows)

    # ── Bulk story add (CSV/XLSX import → review prefill) ────────────────────

    @app.route("/project/<int:project_id>/bulk-add/csv", methods=["POST"])
    @super_user_required
    def bulk_add_csv(project_id):
        project = _check_project_access(project_id)
        enforce_csrf()

        f = request.files.get("import_file") or request.files.get("csv_file")
        if not f or not f.filename:
            flash("No file selected.", "error")
            return redirect(url_for("bulk_add", project_id=project_id))

        fname_lower = f.filename.lower()
        is_xlsx = fname_lower.endswith(".xlsx")

        def norm(s): return (s or "").strip().lower()

        raw_rows = []
        header_keys = set()
        try:
            if is_xlsx:
                from openpyxl import load_workbook
                wb = load_workbook(filename=io.BytesIO(f.read()),
                                   read_only=True, data_only=True)
                ws = wb.worksheets[0]
                it = ws.iter_rows(values_only=True)
                header = None
                for raw_row in it:
                    if raw_row is None:
                        continue
                    if not any(c not in (None, "") for c in raw_row):
                        continue
                    header = [norm(str(c) if c is not None else "") for c in raw_row]
                    break
                if not header:
                    flash("XLSX has no header row.", "warning")
                    return redirect(url_for("bulk_add", project_id=project_id))
                header_keys = set(header)
                for raw_row in it:
                    if raw_row is None:
                        continue
                    vals = ["" if c is None else str(c).strip() for c in raw_row]
                    if not any(vals):
                        continue
                    raw_rows.append({header[i]: (vals[i] if i < len(vals) else "")
                                     for i in range(len(header))})
            else:
                raw = f.read().decode("utf-8-sig", errors="replace")
                reader = csv.DictReader(io.StringIO(raw))
                header_keys = set(norm(k) for k in (reader.fieldnames or []))
                for r in reader:
                    raw_rows.append({norm(k): (v or "").strip() for k, v in r.items()})
        except Exception as e:
            flash(f"Could not read file: {escape(str(e))}", "error")
            return redirect(url_for("bulk_add", project_id=project_id))

        if not raw_rows:
            flash("File had no data rows.", "warning")
            return redirect(url_for("bulk_add", project_id=project_id))

        # Header formats supported:
        # v2 canonical: actor,needs,this,where,to,what
        # v1 canonical: actor,want,verb,x,for,y
        # legacy:       actor,verb,z,x,for,y
        if any(k in header_keys for k in ("needs", "this", "where", "to", "what")):
            grammar_mode = "v2"
        elif "want" in header_keys:
            grammar_mode = "v1"
        else:
            grammar_mode = "legacy"

        # Header sanity — report missing required columns and unknown extras.
        known_cols = {
            "actor", "needs", "want", "verb", "this", "z", "action", "action_word",
            "where", "x", "to", "for", "connector", "what", "y", "outcome", "why",
            "points", "priority", "type", "story_type", "epic",
            "assignee", "description", "os", "software_version", "version",
        }
        if grammar_mode == "v2":
            required_cols = ["actor", "needs", "this", "where", "to", "what"]
        elif grammar_mode == "v1":
            required_cols = ["actor", "want", "verb", "x", "for", "y"]
        else:
            required_cols = ["actor", "verb", "z", "x", "for", "y"]
        missing_required = [c for c in required_cols if c not in header_keys]
        unknown_cols = sorted(h for h in header_keys if h and h not in known_cols)

        # Resolve human-friendly columns to ids for the prefilled selects.
        users_map = {u["display_name"].lower(): u["id"]
                     for u in get_all_active_users()}
        types_map = {t["name"].lower(): t["id"]
                     for t in get_story_types(project_id)}
        epics_map = {e["title"].lower(): e["id"]
                     for e in get_epics(project_id)}

        valid_verbs  = {v.lower(): v for v in VERB_OPTIONS}
        valid_conns  = {c.lower(): c for c in CONNECTOR_OPTIONS}
        valid_actors = {a.lower(): a for a in ACTOR_OPTIONS}

        FIB = {0, 1, 2, 3, 5, 8, 13, 21}

        prefill = []
        unresolved_notes = []  # list[(row_num, [messages])]
        blank_skipped = 0
        # Per-column "rescue" counters for the summary report.
        rescued = {
            "needs": 0, "to": 0, "points": 0, "priority": 0,
            "type": 0, "epic": 0, "assignee": 0,
        }
        explicit_none = {"type": 0, "epic": 0, "assignee": 0}

        for idx, r in enumerate(raw_rows, start=1):
            actor    = r.get("actor", "")
            if grammar_mode == "v2":
                verb = r.get("needs", "") or r.get("want", "") or r.get("verb", "")
                z    = r.get("this", "") or r.get("action_word", "") or r.get("action", "")
                x    = r.get("where", "") or r.get("x", "")
                for_conn = r.get("to", "") or r.get("for", "") or r.get("connector", "")
                y    = r.get("what", "") or r.get("y", "") or r.get("outcome", "") or r.get("why", "")
            elif grammar_mode == "v1":
                verb = r.get("want", "")
                z    = r.get("verb", "") or r.get("action_word", "")
                x    = r.get("x", "") or r.get("where", "")
                for_conn = r.get("for", "") or r.get("to", "") or r.get("connector", "")
                y        = r.get("y", "") or r.get("what", "") or r.get("outcome", "") or r.get("why", "")
            else:
                verb = r.get("verb", "")
                z    = r.get("z", "") or r.get("action", "") or r.get("action_word", "")
                x    = r.get("x", "") or r.get("where", "") or r.get("what", "")
                for_conn = r.get("for", "") or r.get("to", "") or r.get("connector", "")
                y        = r.get("y", "") or r.get("outcome", "") or r.get("why", "")

            if not any((actor, verb, z, x, for_conn, y)):
                blank_skipped += 1
                continue

            notes = []

            # Snap dropdown values to canonical casing where we can.
            if actor and actor.lower() in valid_actors:
                actor = valid_actors[actor.lower()]
            if verb and verb.lower() in valid_verbs:
                verb = valid_verbs[verb.lower()]
            elif verb:
                notes.append(f"needs \"{verb}\" not in allowed list — please pick one.")
                verb = ""
                rescued["needs"] += 1
            if for_conn and for_conn.lower() in valid_conns:
                for_conn = valid_conns[for_conn.lower()]
            elif for_conn:
                notes.append(f"to \"{for_conn}\" not in allowed list — please pick one.")
                for_conn = ""
                rescued["to"] += 1

            # Points → keep only if Fibonacci, else clear with a note.
            pts_raw = (r.get("points", "") or "").strip()
            pts = None
            if pts_raw:
                try:
                    pv = int(pts_raw)
                    if pv in FIB:
                        pts = pv
                    else:
                        notes.append(
                            f"points \"{pts_raw}\" is not Fibonacci — cleared.")
                        rescued["points"] += 1
                except ValueError:
                    notes.append(f"points \"{pts_raw}\" is not a number — cleared.")
                    rescued["points"] += 1

            # Priority → uppercase if valid, else clear with a note.
            prio_raw = (r.get("priority", "") or "").strip().upper()
            prio = prio_raw if prio_raw in ("VH","H","M","L","VL") else ""
            if prio_raw and not prio:
                notes.append(f"priority \"{prio_raw}\" not allowed — cleared.")
                rescued["priority"] += 1

            # Type / epic / assignee → resolve names to ids; note misses.
            # Literal "(none)" (from the XLSX dropdown) is an intentional blank.
            def _is_none_marker(s):
                return (s or "").strip().lower() in ("(none)", "none", "-", "—")

            t_raw_disp = r.get("type", "") or r.get("story_type", "")
            if _is_none_marker(t_raw_disp):
                t_raw_disp, st_id = "", None
                explicit_none["type"] += 1
            else:
                st_id = types_map.get(t_raw_disp.lower()) if t_raw_disp else None
                if t_raw_disp and st_id is None:
                    notes.append(f"type \"{t_raw_disp}\" not found — please pick one.")
                    rescued["type"] += 1
            e_raw_disp = r.get("epic", "")
            if _is_none_marker(e_raw_disp):
                e_raw_disp, ep_id = "", None
                explicit_none["epic"] += 1
            else:
                ep_id = epics_map.get(e_raw_disp.lower()) if e_raw_disp else None
                if e_raw_disp and ep_id is None:
                    notes.append(f"epic \"{e_raw_disp}\" not found — please pick one.")
                    rescued["epic"] += 1
            a_raw_disp = r.get("assignee", "")
            if _is_none_marker(a_raw_disp):
                a_raw_disp, a_id = "", None
                explicit_none["assignee"] += 1
            else:
                a_id = users_map.get(a_raw_disp.lower()) if a_raw_disp else None
                if a_raw_disp and a_id is None:
                    notes.append(f"assignee \"{a_raw_disp}\" not found — please pick one.")
                    rescued["assignee"] += 1

            prefill.append({
                "actor":        actor,
                "verb":         verb,        # needs (column 2 select)
                "z":            z,           # this (column 3 input)
                "x":            x,
                "for_conn":     for_conn,
                "y":            y,
                "points":       pts,
                "priority":     prio,
                "story_type_id": st_id,
                "epic_id":      ep_id,
                "assignee_id":  a_id,
                "description":  r.get("description", ""),
            })
            if notes:
                unresolved_notes.append((idx, notes))

        if not prefill:
            flash("File had no usable rows.", "warning")
            return redirect(url_for("bulk_add", project_id=project_id))

        session["bulk_prefill"] = prefill
        kind_label = "XLSX" if is_xlsx else "CSV"

        # ── Always-on import report ────────────────────────────────────────
        total = len(raw_rows)
        loaded = len(prefill)
        flagged = len(unresolved_notes)
        clean = loaded - flagged

        summary = [
            "<strong>📋 {} import report — {} of {} row{} loaded.</strong>"
            .format(kind_label, loaded, total, "" if total == 1 else "s"),
            "<br>• {} clean • {} flagged • {} blank skipped"
            .format(clean, flagged, blank_skipped),
        ]
        if missing_required:
            summary.append(
                "<br>⚠ Missing required column{}: <code>{}</code> — those fields will be blank."
                .format("" if len(missing_required) == 1 else "s",
                        str(escape(", ".join(missing_required)))))
        if unknown_cols:
            summary.append(
                "<br>ℹ Ignored unknown column{}: <code>{}</code>"
                .format("" if len(unknown_cols) == 1 else "s",
                        str(escape(", ".join(unknown_cols)))))
        rescue_bits = [f"{k}: {v}" for k, v in rescued.items() if v]
        if rescue_bits:
            summary.append("<br>🔧 Cleared invalid values — "
                           + str(escape(", ".join(rescue_bits))))
        none_bits = [f"{k}: {v}" for k, v in explicit_none.items() if v]
        if none_bits:
            summary.append("<br>∅ Explicit (none) — "
                           + str(escape(", ".join(none_bits))))
        summary.append(
            "<br>Review the rows below, fix anything flagged, then click "
            "<strong>Create all valid stories</strong>.")
        flash(Markup("".join(summary)),
              "warning" if (flagged or missing_required) else "success")

        if unresolved_notes:
            lines = ["<strong>Rows needing attention:</strong>"]
            for n, msgs in unresolved_notes:
                lines.append("<br><strong>Row {}:</strong>".format(n))
                for m in msgs:
                    lines.append("<br>&nbsp;&nbsp;• " + str(escape(m)))
            flash(Markup("".join(lines)), "warning")

        return redirect(url_for("bulk_add", project_id=project_id))

    @app.route("/project/<int:project_id>/bulk-add/csv-template")
    @super_user_required
    def bulk_add_csv_template(project_id):
        _check_project_access(project_id)
        sample = (
            "actor,needs,this,where,to,what,points,priority,type,epic,assignee,description,os,software_version\n"
            "User,needs,save progress,Main Menu Demo Version,to,choose what to do later,,,Feature,,,Add a manual save button to the pause menu.,,\n"
            "User,needs,fight tuned enemies,Combat Sandbox,for,balanced game combat,,,Task,,,,,\n"
        )
        # Note: leave points & priority BLANK on bulk import — the team
        # sizes and prioritises stories together inside Grooming
        # (collaborative real-time refinement). Importing guesses
        # poisons the conversation before it starts.
        from flask import Response
        return Response(
            sample,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=gyra-bulk-template.csv"},
        )

    @app.route("/project/<int:project_id>/bulk-add/xlsx-template")
    @super_user_required
    def bulk_add_xlsx_template(project_id):
        _check_project_access(project_id)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.comments import Comment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.utils import get_column_letter
        from routes.helpers import ACTOR_OPTIONS, VERB_OPTIONS, CONNECTOR_OPTIONS
        from flask import Response

        header = ["actor","needs","this","where","to","what","points","priority","type",
                  "epic","assignee","description","os","software_version"]
        sample_rows = [
            ["User","needs","save progress","Main Menu Demo Version","to",
             "choose what to do later","","","Feature","","",
             "Add a manual save button to the pause menu.","",""],
            ["User","needs","fight tuned enemies","Combat Sandbox","for",
             "balanced game combat","","","Task","","","","",""],
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Stories"
        ws.append(header)

        # ── Group-coloured headers (grammar / estimation / classification / meta)
        GROUP = {
            "actor":"grammar","needs":"grammar","this":"grammar",
            "where":"grammar","to":"grammar","what":"grammar",
            "points":"est","priority":"est",
            "type":"class","epic":"class","assignee":"class",
            "description":"meta","os":"meta","software_version":"meta",
        }
        GROUP_FILL = {
            "grammar":   PatternFill("solid", fgColor="1E3A8A"),  # indigo
            "est":       PatternFill("solid", fgColor="B45309"),  # amber-700
            "class":     PatternFill("solid", fgColor="6D28D9"),  # violet-700
            "meta":      PatternFill("solid", fgColor="334155"),  # slate-700
        }
        REQUIRED = {"actor","needs","this","where","to","what"}

        head_font_req = Font(bold=True, color="FFFFFF", size=12)
        head_font_opt = Font(bold=True, color="E5E7EB", italic=True, size=11)
        thin = Side(style="thin", color="111111")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        for col_idx, name in enumerate(header, start=1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = GROUP_FILL[GROUP[name]]
            c.font = head_font_req if name in REQUIRED else head_font_opt
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[1].height = 26

        # ── Header comments to disambiguate and teach
        ws.cell(row=1, column=header.index("needs") + 1).comment = Comment(
            "Need phrase: " + ", ".join(VERB_OPTIONS), "GYRA")
        ws.cell(row=1, column=header.index("this") + 1).comment = Comment(
            "A single observable action (verb): crawl, press button, save, "
            "jump. Start with the verb, not a noun phrase.",
            "GYRA")
        ws.cell(row=1, column=header.index("to") + 1).comment = Comment(
            "Connector: " + ", ".join(CONNECTOR_OPTIONS), "GYRA")
        ws.cell(row=1, column=header.index("points") + 1).comment = Comment(
            "LEAVE BLANK on import.\n\n"
            "Story points are SIZES (compared to each other), not hours. "
            "The team sizes stories TOGETHER inside GYRA's Grooming view "
            "— our real-time collaborative refinement tool. Everyone votes "
            "at the same time, disagreements surface hidden assumptions, "
            "and the number comes out honest.\n\n"
            "If you must fill it: Fibonacci only (0, 1, 2, 3, 5, 8, 13, 21). "
            "Anything over 8 will be flagged as too big.",
            "GYRA")
        ws.cell(row=1, column=header.index("priority") + 1).comment = Comment(
            "LEAVE BLANK on import.\n\n"
            "Priority is set during Grooming — the Product Owner decides "
            "the order, the team sees the trade-offs together. Importing "
            "a guess just creates noise you have to undo later.\n\n"
            "If you must fill it: VH, H, M, L, VL.",
            "GYRA")
        ws.cell(row=1, column=header.index("what") + 1).comment = Comment(
            "Outcome: what result does the Actor need?", "GYRA")

        # ── Sample rows
        for r in sample_rows:
            ws.append(r)

        # ── Column widths + frozen header
        widths = {"actor":13,"needs":16,"this":26,"where":30,"to":12,"what":36,
                  "points":8,"priority":10,"type":15,"epic":20,"assignee":18,
                  "description":42,"os":12,"software_version":18}
        for i, name in enumerate(header, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)
        ws.freeze_panes = "A2"

        DATA_ROWS = 1000
        last_row = DATA_ROWS + 1  # incl. header

        # ── Default wrap-text on every data cell so nothing overflows
        wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
        for row_i in range(2, last_row + 1):
            for col_i in range(1, len(header) + 1):
                ws.cell(row=row_i, column=col_i).alignment = wrap_align
        # Header keeps centered alignment but also wraps long labels
        for col_i in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col_i)
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center")

        # ── Lists sheet (kept visible — OnlyOffice refuses DV refs to hidden sheets)
        lists_ws = wb.create_sheet(title="Lists")

        priorities  = ["VH","H","M","L","VL"]
        points_vals = [0,1,2,3,5,8,13,21]
        types_rows  = list(get_story_types(project_id))
        types_list  = ["(none)"] + [t["name"] for t in types_rows]
        epics_rows  = list(get_epics(project_id))
        epics_list  = ["(none)"] + [e["title"] for e in epics_rows]
        users_list  = ["(none)"] + [u["display_name"] for u in get_all_active_users()]

        list_specs = [
            ("actor",     list(ACTOR_OPTIONS)),
            ("needs",     list(VERB_OPTIONS)),
            ("to",        list(CONNECTOR_OPTIONS)),
            ("points",    points_vals),
            ("priority",  priorities),
            ("type",      types_list),
            ("epic",      epics_list),
            ("assignee",  users_list),
        ]

        col_ranges = {}
        for ci, (col_name, values) in enumerate(list_specs, start=1):
            letter = get_column_letter(ci)
            lists_ws.cell(row=1, column=ci, value=col_name).font = Font(bold=True)
            for ri, v in enumerate(values, start=2):
                lists_ws.cell(row=ri, column=ci, value=v)
            if values:
                col_ranges[col_name] = f"Lists!${letter}$2:${letter}${len(values)+1}"

        # ── Strict DataValidations on the main sheet
        for col_name, formula in col_ranges.items():
            try:
                idx = header.index(col_name) + 1
            except ValueError:
                continue
            target_letter = get_column_letter(idx)
            # epic/assignee can be blank (optional + may be empty list);
            # everything else: strict allow_blank for skipped rows but
            # rejects free text via showErrorMessage.
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle=f"Invalid {col_name}",
                error=f"Pick a value from the {col_name} dropdown.",
                promptTitle=col_name.upper(),
                prompt=f"Pick a {col_name} from the list.",
                showInputMessage=False,
            )
            ws.add_data_validation(dv)
            dv.add(f"{target_letter}2:{target_letter}{last_row}")

        # ── Conditional formatting: priority colour-codes the row
        PRIO_FILLS = {
            "VH": ("DC2626", "FFFFFF"),
            "H":  ("EA580C", "FFFFFF"),
            "M":  ("CA8A04", "111111"),
            "L":  ("16A34A", "FFFFFF"),
            "VL": ("64748B", "FFFFFF"),
        }
        prio_letter = get_column_letter(header.index("priority") + 1)
        prio_cell_range = f"{prio_letter}2:{prio_letter}{last_row}"
        for level, (bg, fg) in PRIO_FILLS.items():
            ws.conditional_formatting.add(
                prio_cell_range,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{level}"'],
                    fill=PatternFill("solid", fgColor=bg),
                    font=Font(color=fg, bold=True),
                ),
            )
        # Row-level tint for priority: paint the what/outcome cell so the row
        # glances coloured without overpowering the cells.
        y_letter = get_column_letter(header.index("what") + 1)
        y_range = f"{y_letter}2:{y_letter}{last_row}"
        for level, (bg, _fg) in PRIO_FILLS.items():
            ws.conditional_formatting.add(
                y_range,
                FormulaRule(
                    formula=[f'=UPPER(${prio_letter}2)="{level}"'],
                    fill=PatternFill("solid", fgColor=bg),
                    font=Font(color="FFFFFF" if level != "M" else "111111"),
                ),
            )

        # ── Conditional formatting: type cell takes the type's own colour
        type_letter = get_column_letter(header.index("type") + 1)
        type_range = f"{type_letter}2:{type_letter}{last_row}"
        for t in types_rows:
            color = (t["color"] or "").lstrip("#").upper() or "9CA3AF"
            if len(color) == 3:
                color = "".join(ch * 2 for ch in color)
            if len(color) != 6:
                color = "9CA3AF"
            ws.conditional_formatting.add(
                type_range,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{t["name"]}"'],
                    fill=PatternFill("solid", fgColor=color),
                    font=Font(color="FFFFFF", bold=True),
                ),
            )

        # ── Conditional formatting: epic cell takes the epic's own colour
        epic_letter = get_column_letter(header.index("epic") + 1)
        epic_range = f"{epic_letter}2:{epic_letter}{last_row}"
        for e in epics_rows:
            ecolor = (e["color"] or "").lstrip("#").upper() or "6B7280"
            if len(ecolor) == 3:
                ecolor = "".join(ch * 2 for ch in ecolor)
            if len(ecolor) != 6 or not all(c in "0123456789ABCDEF" for c in ecolor):
                ecolor = "6B7280"
            # YIQ luminance pick: dark bg → white text, light bg → near-black.
            try:
                r_ = int(ecolor[0:2], 16); g_ = int(ecolor[2:4], 16); b_ = int(ecolor[4:6], 16)
                yiq = (r_ * 299 + g_ * 587 + b_ * 114) / 1000
                fg_color = "1A1208" if yiq >= 160 else "FFFFFF"
            except Exception:
                fg_color = "FFFFFF"
            # Escape any double quotes in the epic title for the formula literal.
            title_lit = (e["title"] or "").replace('"', '""')
            ws.conditional_formatting.add(
                epic_range,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{title_lit}"'],
                    fill=PatternFill("solid", fgColor=ecolor),
                    font=Font(color=fg_color, bold=True),
                ),
            )

        # ── Conditional formatting: points heat
        POINTS_HEAT = {
            0:  ("E5E7EB", "111111"),
            1:  ("DCFCE7", "111111"),
            2:  ("BBF7D0", "111111"),
            3:  ("86EFAC", "111111"),
            5:  ("FDE68A", "111111"),
            8:  ("FCA5A5", "111111"),
            13: ("F87171", "FFFFFF"),
            21: ("DC2626", "FFFFFF"),
        }
        pts_letter = get_column_letter(header.index("points") + 1)
        pts_range = f"{pts_letter}2:{pts_letter}{last_row}"
        for v, (bg, fg) in POINTS_HEAT.items():
            ws.conditional_formatting.add(
                pts_range,
                CellIsRule(
                    operator="equal",
                    formula=[str(v)],
                    fill=PatternFill("solid", fgColor=bg),
                    font=Font(color=fg, bold=True),
                ),
            )

        # ── Highlight missing required cells (red tint)
        for col_name in REQUIRED:
            letter = get_column_letter(header.index(col_name) + 1)
            rng = f"{letter}2:{letter}{last_row}"
            ws.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[
                        f'=AND(COUNTA($A2:$N2)>0,TRIM({letter}2)="")'
                    ],
                    fill=PatternFill("solid", fgColor="FEE2E2"),
                    font=Font(color="991B1B", bold=True),
                ),
            )

        # ── Word-count guard (> 19 words anywhere in actor..what → red)
        wc_formula = (
            '=SUMPRODUCT('
            'IF(TRIM($A2:$F2)="",0,'
            'LEN(TRIM($A2:$F2))-LEN(SUBSTITUTE(TRIM($A2:$F2)," ",""))+1'
            '))>19'
        )
        for letter in ["A","B","C","D","E","F"]:
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{last_row}",
                FormulaRule(
                    formula=[wc_formula],
                    fill=PatternFill("solid", fgColor="FECACA"),
                ),
            )

        # ── Alternating zebra striping on non-grammar columns
        zebra_fill = PatternFill("solid", fgColor="F8FAFC")
        for letter in ["G","H","I","J","K","L","M","N"]:
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{last_row}",
                FormulaRule(
                    formula=[f"=MOD(ROW(),2)=0"],
                    fill=zebra_fill,
                ),
            )

        # ── Visible Help sheet
        help_ws = wb.create_sheet(title="Help", index=1)
        help_ws.sheet_view.showGridLines = False
        help_ws.column_dimensions["A"].width = 22
        help_ws.column_dimensions["B"].width = 90

        def H(row, label, value, label_fill="0F172A", value_fill=None):
            a = help_ws.cell(row=row, column=1, value=label)
            a.font = Font(bold=True, color="FFFFFF")
            a.fill = PatternFill("solid", fgColor=label_fill)
            a.alignment = Alignment(vertical="center", horizontal="right")
            b = help_ws.cell(row=row, column=2, value=value)
            b.alignment = Alignment(wrap_text=True, vertical="center")
            if value_fill:
                b.fill = PatternFill("solid", fgColor=value_fill)
                b.font = Font(color="FFFFFF", bold=True)
            help_ws.row_dimensions[row].height = 22

        help_ws.cell(row=1, column=1, value="GYRA Bulk-Add Template").font = Font(
            bold=True, size=16, color="0F172A")
        help_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        help_ws.row_dimensions[1].height = 30

        H(3, "Grammar",
            "[actor] [needs] [this] [where] [to/for] [what]   —   e.g.  "
            "User needs press button Main Menu Demo Version to choose what to do.")
        H(4, "Required",
            ", ".join(sorted(REQUIRED)) + "  (every row must fill these)")
        H(5, "Max words", "19 words total across actor..what. Overflow rows are rejected.")
        H(6, "actor", ", ".join(ACTOR_OPTIONS)
            + "   — actors RECEIVE value. Never write from staff POV "
            "(Designer/Developer/Artist/QA) unless the story IS a tool for them.")
        H(7, "needs",  ", ".join(VERB_OPTIONS))
        H(8, "this",  "A single observable action (verb): crawl, press button, "
            "save, jump. ❌ 'tuning enemy damage'  ✅ 'fight tuned enemies'.")
        H(9, "to",   ", ".join(CONNECTOR_OPTIONS)
            + "   — use 'to' for outcomes, 'for' for purposes/recipients.")
        H(10,"points","LEAVE BLANK — sized together in Grooming (real-time "
            "collaborative refinement). Fibonacci only if you insist: "
            "0, 1, 2, 3, 5, 8, 13, 21.")
        H(11,"priority","LEAVE BLANK — set during Grooming by the Product Owner "
            "with the team. Allowed if you insist: VH, H, M, L, VL.")
        H(12,"type",  ", ".join(types_list) if types_list else "(no types defined for this project)")
        H(13,"epic",  ", ".join(epics_list) if epics_list else "(no epics defined yet — leave blank)")
        H(14,"assignee", ", ".join(users_list) if users_list else "(no active users)")

        # Priority colour legend
        help_ws.cell(row=16, column=1, value="Priority legend").font = Font(
            bold=True, size=12)
        for off, (level, (bg, fg)) in enumerate(PRIO_FILLS.items()):
            cell = help_ws.cell(row=17 + off, column=2, value=f"{level}")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color=fg, bold=True)
            cell.alignment = Alignment(horizontal="center")
            help_ws.cell(row=17 + off, column=1, value="").alignment = Alignment(horizontal="right")

        # Type colour legend
        type_start = 17 + len(PRIO_FILLS) + 1
        help_ws.cell(row=type_start, column=1, value="Type legend").font = Font(
            bold=True, size=12)
        for off, t in enumerate(types_rows):
            color = (t["color"] or "").lstrip("#").upper() or "9CA3AF"
            if len(color) == 3:
                color = "".join(ch * 2 for ch in color)
            if len(color) != 6:
                color = "9CA3AF"
            cell = help_ws.cell(row=type_start + 1 + off, column=2, value=t["name"])
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Points heat legend
        pts_start = type_start + 1 + max(len(types_rows), 1) + 1
        help_ws.cell(row=pts_start, column=1, value="Points heat").font = Font(
            bold=True, size=12)
        for off, (v, (bg, fg)) in enumerate(POINTS_HEAT.items()):
            cell = help_ws.cell(row=pts_start + 1 + off, column=2, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color=fg, bold=True)
            cell.alignment = Alignment(horizontal="center")

        # ── Reference sheets (ignored by the importer; for human reference) ──
        # Helpers for safe row appending: convert sqlite Row to a plain list
        def _row_get(r, key, default=""):
            try:
                v = r[key]
            except (IndexError, KeyError):
                return default
            return default if v is None else v

        def _fmt_ts(ts):
            if not ts:
                return ""
            try:
                from datetime import datetime
                return datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M")
            except Exception:
                return str(ts)

        def _style_ref_header(sheet, headers, fill_hex):
            sheet.append(headers)
            fill = PatternFill("solid", fgColor=fill_hex)
            font = Font(bold=True, color="FFFFFF", size=12)
            for col_i in range(1, len(headers) + 1):
                cc = sheet.cell(row=1, column=col_i)
                cc.fill = fill
                cc.font = font
                cc.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
            sheet.row_dimensions[1].height = 24
            sheet.freeze_panes = "A2"

        # All Epics for this project (ignored by importer)
        # Minimal view: column C only, one coloured cell per epic title.
        epics_ws = wb.create_sheet(title="All Epics")
        epics_ws.column_dimensions[get_column_letter(1)].width = 70
        epics_ws.column_dimensions[get_column_letter(2)].width = 5
        epics_ws.column_dimensions[get_column_letter(3)].width = 52

        def _hex_to_text_color(hexv):
            # YIQ luminance pick: dark bg → white text, light bg → near-black.
            try:
                r = int(hexv[0:2], 16)
                g = int(hexv[2:4], 16)
                b = int(hexv[4:6], 16)
                yiq = (r * 299 + g * 587 + b * 114) / 1000
                return "1A1208" if yiq >= 160 else "FFFFFF"
            except Exception:
                return "FFFFFF"

        all_epics = [e for e in get_epics(project_id, include_archived=False)
                     if not _row_get(e, "is_archived", 0)]
        for idx, e in enumerate(all_epics, start=1):
            # Column B: checkbox (Unicode ballot box, toggleable via DV dropdown)
            cb = epics_ws.cell(row=idx, column=2, value="☐")
            cb.alignment = Alignment(horizontal="center", vertical="center")
            cb.font = Font(name="Arial Unicode MS", size=14, bold=True)

            cell = epics_ws.cell(row=idx, column=3, value=_row_get(e, "title"))
            cval = _row_get(e, "color") or ""
            hexv = str(cval).lstrip("#").upper()
            if len(hexv) == 3:
                hexv = "".join(ch * 2 for ch in hexv)
            if len(hexv) == 6 and all(c in "0123456789ABCDEF" for c in hexv):
                cell.fill = PatternFill("solid", fgColor=hexv)
                cell.font = Font(color=_hex_to_text_color(hexv), bold=True)
            cell.alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            epics_ws.row_dimensions[idx].height = 22
        # Data validation: clickable dropdown ☐ / ☒ in column B
        if all_epics:
            dv = DataValidation(
                type="list", formula1='"☐,☒"', allow_blank=True)
            dv.error = "Pick ☐ or ☒"
            dv.errorTitle = "Invalid"
            dv.add(f"B1:B{len(all_epics)}")
            epics_ws.add_data_validation(dv)
        # Note as a comment on the first epic cell (if any)
        if epics_ws.max_row >= 1:
            epics_ws.cell(row=1, column=3).comment = Comment(
                "Reference only. This sheet is IGNORED by the importer.", "GYRA")

        # All Initiatives for this project (ignored by importer)
        inits_ws = wb.create_sheet(title="All Initiatives")
        init_headers = ["id", "name", "description", "status", "rule_type",
                        "color", "epic_count", "milestone_count",
                        "created_at", "updated_at"]
        _style_ref_header(inits_ws, init_headers, "059669")
        all_inits = get_initiatives(project_id)
        for it in all_inits:
            inits_ws.append([
                _row_get(it, "id"),
                _row_get(it, "name"),
                _row_get(it, "description"),
                _row_get(it, "status"),
                _row_get(it, "rule_type"),
                _row_get(it, "color"),
                _row_get(it, "epic_count", 0),
                _row_get(it, "milestone_count", 0),
                _fmt_ts(_row_get(it, "created_at", 0)),
                _fmt_ts(_row_get(it, "updated_at", 0)),
            ])
        init_widths = [6, 28, 60, 12, 14, 10, 10, 12, 18, 18]
        for i, w in enumerate(init_widths, start=1):
            inits_ws.column_dimensions[get_column_letter(i)].width = w
        for row_i in range(2, inits_ws.max_row + 1):
            for col_i in range(1, len(init_headers) + 1):
                inits_ws.cell(row=row_i, column=col_i).alignment = Alignment(
                    wrap_text=True, vertical="top")
        # Colour the color column
        col_color_idx = init_headers.index("color") + 1
        for row_i in range(2, inits_ws.max_row + 1):
            cval = inits_ws.cell(row=row_i, column=col_color_idx).value or ""
            hexv = str(cval).lstrip("#").upper()
            if len(hexv) == 3:
                hexv = "".join(ch * 2 for ch in hexv)
            if len(hexv) == 6 and all(c in "0123456789ABCDEF" for c in hexv):
                inits_ws.cell(row=row_i, column=col_color_idx).fill = PatternFill(
                    "solid", fgColor=hexv)
                inits_ws.cell(row=row_i, column=col_color_idx).font = Font(
                    color="FFFFFF", bold=True)
        # Status colour-code
        INIT_STATUS_FILL = {
            "active":   ("D1FAE5", "065F46"),
            "draft":    ("E5E7EB", "374151"),
            "shipped":  ("DBEAFE", "1E40AF"),
            "archived": ("E5E7EB", "6B7280"),
        }
        status_idx = init_headers.index("status") + 1
        for row_i in range(2, inits_ws.max_row + 1):
            sval = (inits_ws.cell(row=row_i, column=status_idx).value or "").lower()
            if sval in INIT_STATUS_FILL:
                bg, fg = INIT_STATUS_FILL[sval]
                cc = inits_ws.cell(row=row_i, column=status_idx)
                cc.fill = PatternFill("solid", fgColor=bg)
                cc.font = Font(color=fg, bold=True)
                cc.alignment = Alignment(horizontal="center")
        inits_ws.cell(row=1, column=1).comment = Comment(
            "Reference only. This sheet is IGNORED by the importer.", "GYRA")

        # ── Sheet order: Stories | All Epics | All Initiatives | Help | Lists
        _order = ["Stories", "All Epics", "All Initiatives", "Help", "Lists"]
        try:
            wb._sheets.sort(
                key=lambda s: _order.index(s.title) if s.title in _order else len(_order)
            )
        except Exception:
            pass

        # ── Move Stories tab to be the active one on open
        wb.active = wb.index(ws)

        # ── Force vertical-center on every cell across every sheet,
        #    while preserving each cell's existing alignment attributes.
        from copy import copy as _copy
        for _sh in wb.worksheets:
            for _row in _sh.iter_rows():
                for _c in _row:
                    _al = _copy(_c.alignment) if _c.alignment else Alignment()
                    _al.vertical = "center"
                    _c.alignment = _al

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=gyra-bulk-template.xlsx"},
        )

    # ── Bulk send to sprint ───────────────────────────────────────────────────

    @app.route("/project/<int:project_id>/bulk-sprint", methods=["POST"])
    @login_required
    def bulk_sprint(project_id):
        enforce_csrf()
        _check_project_access(project_id)
        raw_ids = request.form.getlist("story_ids[]")
        story_ids = [int(x) for x in raw_ids if x.isdigit()]
        if not story_ids:
            flash("No stories selected.", "warning")
            return redirect(url_for("backlog", project_id=project_id))

        sprint = get_current_sprint(project_id)
        conn   = get_db()
        first_status = conn.execute(
            "SELECT id FROM statuses WHERE project_id=? ORDER BY order_index LIMIT 1",
            (project_id,),
        ).fetchone()
        status_id = first_status["id"] if first_status else None
        now = int(time.time())

        ph = ",".join("?" * len(story_ids))
        conn.execute(
            f"UPDATE stories SET sprint=?, status_id=?, updated_at=?"
            f" WHERE id IN ({ph}) AND project_id=? AND sprint IS NULL AND is_archived=0",
            [sprint, status_id, now] + story_ids + [project_id],
        )
        moved = conn.execute("SELECT changes()").fetchone()[0]
        conn.commit()
        conn.close()
        flash(
            f"{moved} {'story' if moved == 1 else 'stories'} added to Sprint {sprint}.",
            "success",
        )
        return redirect(url_for("backlog", project_id=project_id))

    # ── End sprint ────────────────────────────────────────────────────────────

    @app.route("/project/<int:project_id>/end-sprint", methods=["POST"])
    @login_required
    def end_sprint(project_id):
        enforce_csrf()
        if session.get("role") not in ("admin", "super_user"):
            abort(403)
        project = _check_project_access(project_id)  # noqa: assigned for clarity
        sprint  = get_current_sprint(project_id)
        now     = int(time.time())
        conn    = get_db()

        # Archive stories in "done" columns
        conn.execute(
            """UPDATE stories SET is_archived=1, updated_at=?
               WHERE project_id=? AND sprint=?
                 AND status_id IN (SELECT id FROM statuses WHERE project_id=? AND is_done=1)""",
            (now, project_id, sprint, project_id),
        )
        archived = conn.execute("SELECT changes()").fetchone()[0]

        # Return unfinished stories to backlog
        conn.execute(
            """UPDATE stories SET sprint=NULL, updated_at=?
               WHERE project_id=? AND sprint=? AND is_archived=0""",
            (now, project_id, sprint),
        )
        returned = conn.execute("SELECT changes()").fetchone()[0]

        conn.commit()
        conn.close()
        flash(
            f"Sprint {sprint} ended — {archived} "
            f"{'story' if archived == 1 else 'stories'} archived, "
            f"{returned} returned to backlog.",
            "success",
        )
        return redirect(url_for("board", project_id=project_id))

    # ── Archive view ──────────────────────────────────────────────────────────

    @app.route("/project/<int:project_id>/archive")
    @login_required
    def archive(project_id):
        project = _check_project_access(project_id)
        return render_template("archive.html", project=project)

    @app.route("/api/project/<int:project_id>/archive")
    @login_required
    def api_archive(project_id):
        _check_project_access(project_id)
        page     = max(0, int(request.args.get("page", 0)))
        per_page = min(int(request.args.get("per_page", 20)), 50)
        offset   = page * per_page

        conn = get_db()
        rows = conn.execute(
            """SELECT s.id, s.title, s.story_z, s.story_points, s.priority, s.sprint,
                      s.updated_at, s.created_at,
                      st.name AS status_name, st.color AS status_color,
                      sty.name AS story_type_name, sty.color AS story_type_color,
                      p.key AS project_key
               FROM stories s
               LEFT JOIN statuses st ON s.status_id = st.id
               LEFT JOIN story_types sty ON s.story_type = sty.id
               LEFT JOIN projects p ON s.project_id = p.id
               WHERE s.project_id=? AND s.is_archived=1
               ORDER BY s.updated_at DESC
               LIMIT ? OFFSET ?""",
            (project_id, per_page + 1, offset),
        ).fetchall()

        has_more = len(rows) > per_page
        items = [dict(r) for r in rows[:per_page]]

        if items:
            story_ids = [it["id"] for it in items]
            ph = ",".join("?" * len(story_ids))

            sk_rows = conn.execute(
                f"SELECT id, type, card_x, card_y, label, card_story_id"
                f" FROM stickers WHERE card_story_id IN ({ph})",
                story_ids,
            ).fetchall()
            stickers_map: dict = {}
            for sk in sk_rows:
                stickers_map.setdefault(sk["card_story_id"], []).append(dict(sk))

            a_rows = conn.execute(
                f"""SELECT su.story_id, u.display_name, u.avatar
                    FROM story_users su JOIN users u ON su.user_id = u.id
                    WHERE su.story_id IN ({ph})""",
                story_ids,
            ).fetchall()
            assignees_map: dict = {}
            for a in a_rows:
                assignees_map.setdefault(a["story_id"], []).append(
                    {"display_name": a["display_name"], "avatar": a["avatar"]}
                )

            previews_map = get_story_previews(story_ids)

            for it in items:
                it["stickers"]   = stickers_map.get(it["id"], [])
                it["assignees"]  = assignees_map.get(it["id"], [])
                it["images"]     = previews_map.get(it["id"], [])
                it["thumbnail"]  = it["images"][0] if it["images"] else None
                it["html_title"] = str(bold_verb_in_title(it["title"], it.get("story_z") or ""))

        conn.close()
        return jsonify(items=items, has_more=has_more, page=page)

    # ── Unarchive (re-add to active sprint) ───────────────────────────────────

    @app.route("/api/story/<int:story_id>/unarchive", methods=["POST"])
    @login_required
    def api_unarchive_story(story_id):
        enforce_csrf()
        conn  = get_db()
        story = conn.execute("SELECT * FROM stories WHERE id=?", (story_id,)).fetchone()
        if not story:
            conn.close()
            abort(404)
        project_id = story["project_id"]
        if session.get("role") != "admin" and not user_in_project(session["user_id"], project_id):
            conn.close()
            abort(403)

        sprint = get_current_sprint(project_id)
        first_status = conn.execute(
            "SELECT id FROM statuses WHERE project_id=? ORDER BY order_index LIMIT 1",
            (project_id,),
        ).fetchone()
        status_id = first_status["id"] if first_status else story["status_id"]

        conn.execute(
            "UPDATE stories SET is_archived=0, sprint=?, status_id=?, updated_at=? WHERE id=?",
            (sprint, status_id, int(time.time()), story_id),
        )
        conn.commit()
        conn.close()
        return jsonify(ok=True, sprint=sprint)
